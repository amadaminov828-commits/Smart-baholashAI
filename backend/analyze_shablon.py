import pandas as pd
import openpyxl
import json

file_path = r"c:\Users\Asus\Desktop\antigravity\backend\Шаблон недвижимость1121.xlsx"

def analyze_excel(path):
    # Load workbook to see sheets and potentially formulas
    wb = openpyxl.load_workbook(path, data_only=False)
    analysis = {
        "sheets": wb.sheetnames,
        "details": {}
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        analysis["details"][sheet_name] = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "samples": []
        }
        
        # Read first 50 lines to get a sense of structure
        for row in range(1, min(50, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(10, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                # If it's a formula, we want to know
                is_formula = str(val).startswith('=') if val else False
                row_data.append({"val": str(val), "is_formula": is_formula})
            analysis["details"][sheet_name]["samples"].append(row_data)
            
    print(json.dumps(analysis, indent=2, ensure_ascii=False))

if __name__ == "__main__":
    analyze_excel(file_path)
